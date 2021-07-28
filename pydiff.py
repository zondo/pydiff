"""
An extended python diff.

TODO: add metaclass to create the class list for text_source()
"""

import difflib
import warnings


class TextSource(object):
    "Base class of sources of text lines."

    # Name of the source file.
    filename = None

    # Tuple of file suffixes used.  Must contain the period.
    suffixes = ()

    def __init__(self, path):
        raise NotImplementedError

    def lines(self):
        "Yield text lines."
        raise NotImplementedError

    def __repr__(self):
        return "<{}: {}>".format(self.__class__.__name__, self.filename)

    def __str__(self):
        return self.filename


class Spreadsheet(object):
    sheetformat = "=== Sheet: %s ==="
    separator = " | "
    splitnewlines = False
    numformat = "%g"

    def cell_lines(self, cellrow, celltotext):
        values = map(celltotext, cellrow)
        values = filter(lambda x: x, values)
        line = self.separator.join(values).rstrip()

        if self.splitnewlines:
            for string in line.split("\n"):
                yield string
        else:
            yield line.replace("\n", " ")


class Excel(TextSource, Spreadsheet):
    suffixes = (".xls",)

    def __init__(self, path):
        from xlrd import open_workbook
        self.book = open_workbook(path)

    def lines(self):
        from xlrd import XL_CELL_NUMBER

        def totext(cell):
            if cell.ctype == XL_CELL_NUMBER:
                return self.numformat % cell.value
            else:
                return str(cell.value)

        for idx in range(self.book.nsheets):
            sheet = self.book.sheet_by_index(idx)
            yield self.sheetformat % sheet.name

            for row in range(sheet.nrows):
                cells = sheet.row_slice(row)
                yield from self.cell_lines(cells, totext)


class Excel2007(TextSource, Spreadsheet):
    suffixes = (".xlsx", ".xlsm")

    def __init__(self, path):
        from openpyxl import load_workbook

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            self.book = load_workbook(path, data_only=True)

    def lines(self):
        def totext(cell):
            return str(cell.value or "")

        for name in self.book.sheetnames:
            sheet = self.book[name]
            yield self.sheetformat % name

            for cells in sheet.rows:
                yield from self.cell_lines(cells, totext)


def diff_lines(src1, src2, context=False, ignoreblank=True, contextlines=3):
    "Yield diff lines twixt two text sources."

    lines1 = src1.lines()
    lines2 = src2.lines()

    if ignoreblank:
        lines1 = (line for line in lines1 if line)
        lines2 = (line for line in lines2 if line)

    if context:
        func = difflib.context_diff
    else:
        func = difflib.unified_diff

    return func(list(lines1), list(lines2), str(src1), str(src2),
                lineterm="", n=contextlines)


def text_source(path):
    "Return TextSource object for a path."

    for cls in Excel, Excel2007:
        if path.endswith(cls.suffixes):
            src = cls(path)
            src.filename = path
            return src

    raise RuntimeError("no reader found for '%s'" % path)


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 2:
        # Two args -- show diff.
        file1, file2 = sys.argv[1:3]
        src1 = text_source(file1)
        src2 = text_source(file2)
        output = diff_lines(src1, src2, context=True)
    elif len(sys.argv) == 2:
        # One arg -- show text contents.
        path = sys.argv[1]
        output = text_source(path).lines()
    else:
        sys.exit("Usage: pydiff FILE1 [FILE2]")

    for line in output:
        print(line)
